import requests
from bs4 import BeautifulSoup
import openpyxl

# Load the spreadsheet
wb = openpyxl.load_workbook('data.xlsx')
sheet = wb.active

# Process through the tickers in column A
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    ticker = row[0].value
    if ticker:
        url = f'https://stockanalysis.com/stocks/{ticker}/company/'
        response = requests.get(url)

        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'lxml')
            company_info = soup.select_one('#main > div:nth-of-type(2) > div:nth-of-type(1)')
            if company_info:
                paragraphs = company_info.find_all('p')
                full_text = "\n".join(p.get_text() for p in paragraphs)

                sheet[f'B{row[0].row}'] = full_text
            else:
                sheet[f'B{row[0].row}'] = 'Data not found'
        else:
            sheet[f'B{row[0].row}'] = 'Request failed'

# Save the workbook
wb.save('data.xlsx')
